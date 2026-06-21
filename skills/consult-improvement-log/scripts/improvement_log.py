#!/usr/bin/env python3
"""improvement_log.py — unified Item Register sync utility.

JSON source-of-truth <-> Excel/CSV sync for the engagement Item Register.

The register holds every item that hangs off an L2 taxonomy node, discriminated
by `type`:
  - improvement : a process improvement opportunity (Stream B)
  - gap         : a gap / validation item (from the drafter gap tags)
  - screenshot  : a screenshot placeholder (SC-IDs)

Each item links back to its taxonomy node via l1_cycle / l2_process (kebab slugs).

Commands:
  build-xlsx   Build an Excel review workbook from JSON.
  update-json  Merge a CSV review file back into JSON.
  remove       Archive or hard-delete records by ID.
  validate     Check records against the controlled vocab; report issues.

Removal behavior:
  - Safe removal/archive: set record_status=archived, or use remove --archive.
  - Hard delete from CSV: set record_status=deleted_candidate and pass --apply-deletes.
  - Hard delete missing rows: pass --delete-missing to update-json. Use carefully.
  - Hard delete by ID: use remove without --archive.

Validation is non-fatal: unknown/invalid values are flagged (requires_human_review
set True, review_status=needs_review, note appended to change_notes) rather than
rejected, so the register stays extensible.
"""
from __future__ import annotations

import argparse, csv, json, re, shutil
from copy import deepcopy
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

DEFAULT_RECORD_FIELDS = [
    "id", "type", "tag", "dedup_key", "disposition", "date_identified", "source",
    "l1_cycle", "l2_process", "l3_activity",
    "observation_pain_point", "root_cause", "recommended_action",
    "impact_type", "estimated_impact_benefit", "effort", "priority",
    "owner", "phase", "escalation_status", "process_owner_contacts",
    "notes_next_step", "record_status", "review_status", "requires_human_review",
    "last_modified_by", "last_modified_at", "change_notes",
]
PROTECTED_FIELDS = {"id"}
DATE_FIELDS = {"date_identified"}
BOOL_FIELDS = {"requires_human_review"}
DELETE_STATUSES = {"deleted_candidate", "delete", "deleted", "remove", "removed"}
ARCHIVE_STATUSES = {"archived", "inactive"}

# ---- Controlled vocabulary -------------------------------------------------
# Validation is soft: values outside these sets are flagged, not rejected.
ITEM_TYPES = {"improvement", "gap", "screenshot", "unmapped", "theme"}

# For type=improvement, `tag` names the diagnostic lens the item addresses.
LENS_TAGS = {"process", "automation", "operating_model", "capability"}

# For type=gap, `tag` is a normalized gap tag; each maps to a reporting category.
GAP_TAG_CATEGORY = {
    "not_documented": "general", "unconfirmed": "general", "confirm": "general",
    "owner_unknown": "ownership", "reviewer_unknown": "ownership",
    "approver_unknown": "ownership",
    "system_unknown": "process", "timing_unknown": "process",
    "frequency_unknown": "process", "input_unknown": "process",
    "output_unknown": "process", "navigation_unknown": "process",
    "field_unknown": "process",
    "control_not_evidenced": "controls", "approval_not_evidenced": "controls",
    "evidence_retention_unknown": "controls", "archive_location_unknown": "controls",
    "exception_handling_unknown": "exceptions",
    "downstream_dependency_unknown": "exceptions",
    "upstream_dependency_unknown": "exceptions",
}
# Map the drafter's literal bracketed tags onto normalized slugs (for ingestion).
GAP_TAG_ALIASES = {
    "gap — not documented": "not_documented",
    "unconfirmed — validate": "unconfirmed",
    "confirm": "confirm",
    "gap — owner unknown": "owner_unknown",
    "gap — reviewer unknown": "reviewer_unknown",
    "gap — approver unknown": "approver_unknown",
    "gap — system unknown": "system_unknown",
    "gap — timing unknown": "timing_unknown",
    "gap — frequency unknown": "frequency_unknown",
    "gap — input unknown": "input_unknown",
    "gap — output unknown": "output_unknown",
    "gap — navigation path unknown": "navigation_unknown",
    "gap — field / parameter unknown": "field_unknown",
    "gap — control not evidenced": "control_not_evidenced",
    "gap — approval not evidenced": "approval_not_evidenced",
    "gap — evidence retention unknown": "evidence_retention_unknown",
    "gap — archive location unknown": "archive_location_unknown",
    "gap — exception handling unknown": "exception_handling_unknown",
    "gap — downstream dependency unknown": "downstream_dependency_unknown",
    "gap — upstream dependency unknown": "upstream_dependency_unknown",
}
IMPACT_TYPES = {"cost", "time", "risk", "quality", "control"}
EFFORT_LEVELS = {"low", "med", "high"}
PRIORITY_LEVELS = {"p1", "p2", "p3"}


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def normalize_tag(value: Any) -> Any:
    """Normalize a raw tag (possibly a bracketed drafter tag) to a slug."""
    if value is None:
        return None
    txt = str(value).strip().strip("[]").strip().lower()
    if not txt:
        return None
    if txt in GAP_TAG_ALIASES:
        return GAP_TAG_ALIASES[txt]
    return re.sub(r"[^0-9a-z]+", "_", txt).strip("_")


def normalize_header(header: str) -> str:
    h = str(header or "").strip().replace("﻿", "")
    mapping = {
        "ID": "id", "Type": "type", "Tag": "tag",
        "Date Identified": "date_identified", "Source": "source",
        "L1 Cycle": "l1_cycle", "L2 Process": "l2_process", "L3 Activity": "l3_activity",
        "Observation / Pain Point": "observation_pain_point", "Root Cause": "root_cause",
        "Recommended Action": "recommended_action", "Impact Type": "impact_type",
        "Est. Impact / Benefit": "estimated_impact_benefit", "Effort": "effort",
        "Priority": "priority", "Owner": "owner", "Phase": "phase",
        "Escalation Status": "escalation_status",
        "Process Owner Contact(s)": "process_owner_contacts",
        "Notes / Next Step": "notes_next_step",
    }
    if h in mapping:
        return mapping[h]
    h = h.replace("/", " ").replace("&", " and ")
    return re.sub(r"[^0-9A-Za-z]+", "_", h).strip("_").lower()


def validate_record(rec: Dict[str, Any]) -> List[str]:
    """Return a list of human-readable vocab issues for one record (non-fatal)."""
    issues: List[str] = []
    rtype = str(rec.get("type") or "improvement").lower()
    if rtype not in ITEM_TYPES:
        issues.append(f"unknown type '{rtype}'")
    tag = rec.get("tag")
    if tag:
        tag = str(tag).lower()
        if rtype == "improvement" and tag not in LENS_TAGS:
            issues.append(f"improvement tag '{tag}' not a known lens")
        elif rtype == "gap" and tag not in GAP_TAG_CATEGORY:
            issues.append(f"gap tag '{tag}' not a known gap tag")
    for field, allowed in (("impact_type", IMPACT_TYPES), ("effort", EFFORT_LEVELS),
                           ("priority", PRIORITY_LEVELS)):
        val = rec.get(field)
        if val and str(val).lower() not in allowed:
            issues.append(f"{field} '{val}' not in {sorted(allowed)}")
    return issues


def load_source_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, dict) or not isinstance(data.get("records"), list):
        raise ValueError("JSON must be an object with a top-level records list.")
    ids = [r.get("id") for r in data["records"]]
    if any(not x for x in ids):
        raise ValueError("Every record must have a non-empty id.")
    dupes = sorted({x for x in ids if ids.count(x) > 1})
    if dupes:
        raise ValueError(f"Duplicate IDs in JSON: {dupes}")
    return data


def field_order(data: Dict[str, Any]) -> List[str]:
    schema_fields = list((data.get("schema", {}).get("fields", {}) or {}).keys())
    record_fields = []
    for rec in data.get("records", []):
        for key in rec:
            if key not in record_fields:
                record_fields.append(key)
    out = []
    for key in DEFAULT_RECORD_FIELDS + schema_fields + record_fields:
        if key not in out:
            out.append(key)
    return out


def clean_value(value: Any, field: str) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
    if field in BOOL_FIELDS:
        txt = str(value).strip().lower()
        if txt in {"true", "yes", "y", "1"}: return True
        if txt in {"false", "no", "n", "0"}: return False
        return value
    if field == "tag":
        return normalize_tag(value)
    if field in DATE_FIELDS:
        dt = pd.to_datetime(value, errors="coerce")
        if not pd.isna(dt):
            return dt.date().isoformat()
    return value


def save_json(data: Dict[str, Any], json_path: Path, out_path: Path, backup: bool = True) -> None:
    data.setdefault("metadata", {})["record_count"] = len(data["records"])
    if backup and out_path.resolve() == json_path.resolve():
        backup_path = json_path.with_suffix(f".backup-{datetime.now().strftime('%Y%m%d-%H%M%S')}.json")
        shutil.copy2(json_path, backup_path)
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    check = load_source_json(out_path)
    if check.get("metadata", {}).get("record_count") != len(check["records"]):
        raise ValueError("Validation failed: metadata.record_count does not match records length.")


def build_xlsx(json_path: Path, xlsx_path: Path, active_only: bool = False) -> None:
    data = load_source_json(json_path)
    records = data["records"]
    if active_only:
        records = [r for r in records if str(r.get("record_status", "active")).lower() not in ARCHIVE_STATUSES]
    fields = field_order(data)
    pd.DataFrame([{f: r.get(f) for f in fields} for r in records], columns=fields).to_excel(
        xlsx_path, index=False, sheet_name="Item Register", engine="openpyxl"
    )
    wb = load_workbook(xlsx_path)
    ws = wb["Item Register"]
    ws.freeze_panes = "A2"
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True); cell.fill = fill; cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    if ws.max_row >= 2:
        ref = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
        tab = Table(displayName="ItemRegister", ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False, showColumnStripes=False)
        ws.add_table(tab)
    widths = {"id":12, "type":13, "tag":24, "date_identified":15, "source":20,
              "l1_cycle":22, "l2_process":26, "l3_activity":26,
              "observation_pain_point":60, "root_cause":50, "recommended_action":60,
              "estimated_impact_benefit":42, "notes_next_step":60, "change_notes":40}
    for i, f in enumerate(fields, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = widths.get(f, 22)
    wb.save(xlsx_path)


def read_csv_rows(csv_path: Path) -> List[Dict[str, Any]]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError("CSV has no header row.")
        headers = [normalize_header(h) for h in reader.fieldnames]
        rows = []
        for raw in reader:
            row = {h: raw.get(orig) for orig, h in zip(reader.fieldnames, headers) if h}
            if any(v not in (None, "") for v in row.values()):
                rows.append(row)
        return rows


def flag_issues(rec: Dict[str, Any], ts: str) -> None:
    """If a record has vocab issues, flag it for human review (non-fatal)."""
    issues = validate_record(rec)
    if issues:
        rec["requires_human_review"] = True
        rec["review_status"] = "needs_review"
        note = f"[{ts}] vocab check: " + "; ".join(issues)
        existing = str(rec.get("change_notes") or "").strip()
        rec["change_notes"] = (existing + "\n" + note).strip() if existing else note


def update_json(json_path: Path, csv_path: Path, out_path: Path, modified_by: str, apply_deletes: bool, delete_missing: bool) -> Tuple[int,int,int,int,int]:
    data = load_source_json(json_path)
    by_id = {r["id"]: r for r in data["records"]}
    rows = read_csv_rows(csv_path)
    seen, to_delete = set(), set()
    updated = inserted = unchanged = archived = 0
    ts = now_iso()

    for row in rows:
        rec_id = clean_value(row.get("id"), "id")
        if not rec_id:
            raise ValueError("CSV row missing required id.")
        if rec_id in seen:
            raise ValueError(f"Duplicate ID in CSV: {rec_id}")
        seen.add(rec_id)
        clean = {k: clean_value(v, k) for k, v in row.items()}
        status = str(clean.get("record_status") or "").lower()

        if status in DELETE_STATUSES and apply_deletes:
            if rec_id in by_id:
                to_delete.add(rec_id)
            continue

        if rec_id in by_id:
            target = by_id[rec_id]
            before = deepcopy(target)
            for k, v in clean.items():
                if k not in PROTECTED_FIELDS:
                    target[k] = v
            if target != before:
                if status in ARCHIVE_STATUSES:
                    archived += 1
                target["last_modified_by"] = modified_by
                target["last_modified_at"] = ts
                flag_issues(target, ts)
                updated += 1
            else:
                unchanged += 1
        else:
            if status in DELETE_STATUSES:
                unchanged += 1
                continue
            clean.setdefault("type", "improvement")
            clean.setdefault("record_status", "active")
            clean.setdefault("review_status", "needs_review")
            clean.setdefault("requires_human_review", True)
            clean["last_modified_by"] = modified_by
            clean["last_modified_at"] = ts
            flag_issues(clean, ts)
            data["records"].append(clean)
            by_id[rec_id] = clean
            inserted += 1

    if delete_missing:
        to_delete.update(set(by_id) - seen)

    if to_delete:
        data["records"] = [r for r in data["records"] if r.get("id") not in to_delete]

    data.setdefault("metadata", {})["last_updated_at"] = ts
    data["metadata"]["last_update_source"] = str(csv_path)
    data["metadata"]["last_update_deleted_count"] = len(to_delete)
    save_json(data, json_path, out_path)
    return updated, inserted, unchanged, archived, len(to_delete)


def upsert_records(json_path: Path, records: List[Dict[str, Any]], out_path: Path,
                   modified_by: str) -> Tuple[int, int, int]:
    """JSON-native upsert. Upsert a list of record dicts into the register.

    Matching precedence per incoming record:
      1. If the record carries a non-empty `dedup_key`, match an existing record
         with the same `dedup_key` (update in place; else insert).
      2. Otherwise match by `id` (update in place; else insert).

    Unlike the CSV path this does NOT clean/normalize via header maps — it takes
    records as already-structured dicts (only light per-field cleaning is applied
    so booleans/blank strings behave). Returns (updated, inserted, unchanged).
    """
    data = load_source_json(json_path)
    by_id = {r["id"]: r for r in data["records"]}
    # Index existing records by dedup_key (last write wins on collision; the
    # register is validated to have unique ids, dedup_key collisions are merged).
    by_dedup: Dict[str, Dict[str, Any]] = {}
    for r in data["records"]:
        dk = r.get("dedup_key")
        if dk:
            by_dedup[str(dk)] = r

    updated = inserted = unchanged = 0
    ts = now_iso()

    for raw in records:
        if not isinstance(raw, dict):
            raise ValueError("Each upsert record must be an object/dict.")
        clean = {k: clean_value(v, k) for k, v in raw.items()}
        dedup_key = clean.get("dedup_key")
        target = None
        if dedup_key and str(dedup_key) in by_dedup:
            target = by_dedup[str(dedup_key)]
        elif clean.get("id") and clean["id"] in by_id:
            target = by_id[clean["id"]]

        if target is not None:
            before = deepcopy(target)
            for k, v in clean.items():
                if k in PROTECTED_FIELDS:
                    continue  # never overwrite an existing id
                target[k] = v
            if target != before:
                target["last_modified_by"] = modified_by
                target["last_modified_at"] = ts
                flag_issues(target, ts)
                updated += 1
            else:
                unchanged += 1
        else:
            rec_id = clean.get("id")
            if not rec_id:
                raise ValueError("Upsert insert requires a non-empty id when no "
                                 "matching dedup_key exists.")
            if rec_id in by_id:
                raise ValueError(f"Duplicate id on insert: {rec_id}")
            clean.setdefault("type", "improvement")
            clean.setdefault("record_status", "active")
            clean.setdefault("review_status", "needs_review")
            # Honor a caller-provided requires_human_review; default true only
            # when the caller omitted it entirely.
            if "requires_human_review" not in clean or clean.get("requires_human_review") is None:
                clean["requires_human_review"] = True
            clean["last_modified_by"] = modified_by
            clean["last_modified_at"] = ts
            flag_issues(clean, ts)
            data["records"].append(clean)
            by_id[rec_id] = clean
            dk = clean.get("dedup_key")
            if dk:
                by_dedup[str(dk)] = clean
            inserted += 1

    data.setdefault("metadata", {})["last_updated_at"] = ts
    data["metadata"]["last_update_source"] = f"json upsert ({modified_by})"
    save_json(data, json_path, out_path)  # one backup per invocation
    return updated, inserted, unchanged


def remove_records(json_path: Path, ids: List[str], out_path: Path, modified_by: str, archive: bool) -> Tuple[int,int]:
    data = load_source_json(json_path)
    ids = set(ids)
    existing = {r["id"] for r in data["records"]}
    missing = len(ids - existing)
    ts = now_iso()
    if archive:
        affected = 0
        for r in data["records"]:
            if r["id"] in ids:
                r["record_status"] = "archived"
                r["requires_human_review"] = False
                r["last_modified_by"] = modified_by
                r["last_modified_at"] = ts
                affected += 1
    else:
        before = len(data["records"])
        data["records"] = [r for r in data["records"] if r["id"] not in ids]
        affected = before - len(data["records"])
    data.setdefault("metadata", {})["last_updated_at"] = ts
    data["metadata"]["last_update_source"] = "remove command"
    save_json(data, json_path, out_path)
    return affected, missing


def schema_check(data: Dict[str, Any], schema_path: Path) -> List[str]:
    """Validate the register against a JSON Schema. Returns error messages.

    Gracefully degrades if jsonschema or the schema file is absent.
    """
    if not schema_path.exists():
        return [f"(skipped: schema not found at {schema_path})"]
    try:
        import jsonschema
    except ImportError:
        return ["(skipped: jsonschema not installed — pip install jsonschema)"]
    with schema_path.open("r", encoding="utf-8") as f:
        schema = json.load(f)
    validator = jsonschema.Draft7Validator(schema)
    return [f"{'/'.join(str(p) for p in e.path) or '<root>'}: {e.message}"
            for e in sorted(validator.iter_errors(data), key=lambda e: list(e.path))]


def validate_cmd(json_path: Path, schema_path: Path | None = None) -> int:
    data = load_source_json(json_path)
    total_issues = 0
    for rec in data["records"]:
        issues = validate_record(rec)
        if issues:
            total_issues += len(issues)
            print(f"  {rec.get('id')}: " + "; ".join(issues))
    counts: Dict[str, int] = {}
    for rec in data["records"]:
        counts[str(rec.get("type") or "improvement")] = counts.get(str(rec.get("type") or "improvement"), 0) + 1
    print(f"Records: {len(data['records'])} ({', '.join(f'{k}={v}' for k, v in sorted(counts.items()))})")
    print(f"Vocab issues (flagged, non-fatal): {total_issues}")
    if schema_path is not None:
        errors = schema_check(data, schema_path)
        if errors and errors[0].startswith("(skipped"):
            print(f"Schema: {errors[0]}")
        elif errors:
            print(f"Schema: {len(errors)} error(s):")
            for e in errors[:20]:
                print(f"  - {e}")
        else:
            print("Schema: OK.")
    return total_issues


def main() -> None:
    p = argparse.ArgumentParser(description="Sync the Item Register JSON with XLSX/CSV; archive/delete/validate support.")
    sub = p.add_subparsers(dest="command", required=True)
    b = sub.add_parser("build-xlsx")
    b.add_argument("--json", required=True, type=Path); b.add_argument("--xlsx", required=True, type=Path)
    b.add_argument("--active-only", action="store_true", help="Exclude archived/inactive records.")
    u = sub.add_parser("update-json")
    u.add_argument("--json", required=True, type=Path); u.add_argument("--csv", required=True, type=Path); u.add_argument("--out-json", required=True, type=Path)
    u.add_argument("--modified-by", default="csv_update")
    u.add_argument("--apply-deletes", action="store_true", help="Hard-delete rows with record_status=deleted_candidate/delete/deleted/remove/removed.")
    u.add_argument("--delete-missing", action="store_true", help="Hard-delete JSON records absent from the CSV. Use carefully.")
    uj = sub.add_parser("upsert-json", help="JSON-native upsert (by dedup_key, else id). "
                        "Records come from --records-json (inline) or --records-file.")
    uj.add_argument("--json", required=True, type=Path)
    uj.add_argument("--out-json", required=True, type=Path)
    uj.add_argument("--records-json", help="Inline JSON: a list of record dicts (or an "
                    "object with a 'records' list).")
    uj.add_argument("--records-file", type=Path, help="Path to a JSON file holding a list "
                    "of record dicts (or an object with a 'records' list).")
    uj.add_argument("--modified-by", default="json_upsert")
    r = sub.add_parser("remove")
    r.add_argument("--json", required=True, type=Path); r.add_argument("--ids", nargs="+", required=True); r.add_argument("--out-json", required=True, type=Path)
    r.add_argument("--modified-by", default="manual_remove"); r.add_argument("--archive", action="store_true")
    v = sub.add_parser("validate")
    v.add_argument("--json", required=True, type=Path)
    v.add_argument("--schema", type=Path, default=None, help="Optional JSON Schema to validate against.")
    args = p.parse_args()
    if args.command == "build-xlsx":
        build_xlsx(args.json, args.xlsx, args.active_only)
        print(f"Built XLSX: {args.xlsx}")
    elif args.command == "update-json":
        a,b,c,d,e = update_json(args.json, args.csv, args.out_json, args.modified_by, args.apply_deletes, args.delete_missing)
        print(f"Updated JSON: {args.out_json}")
        print(f"Rows updated: {a}; inserted: {b}; unchanged: {c}; archived: {d}; deleted: {e}")
    elif args.command == "upsert-json":
        if bool(args.records_json) == bool(args.records_file):
            raise SystemExit("upsert-json requires exactly one of --records-json / --records-file.")
        if args.records_file:
            with args.records_file.open("r", encoding="utf-8") as f:
                payload = json.load(f)
        else:
            payload = json.loads(args.records_json)
        if isinstance(payload, dict) and isinstance(payload.get("records"), list):
            records = payload["records"]
        elif isinstance(payload, list):
            records = payload
        else:
            raise SystemExit("upsert records must be a JSON list, or an object with a 'records' list.")
        a, b, c = upsert_records(args.json, records, args.out_json, args.modified_by)
        print(f"Upserted JSON: {args.out_json}")
        print(f"Rows updated: {a}; inserted: {b}; unchanged: {c}")
    elif args.command == "remove":
        affected, missing = remove_records(args.json, args.ids, args.out_json, args.modified_by, args.archive)
        print(f"{'Archived' if args.archive else 'Deleted'} records: {affected}; IDs not found: {missing}; output: {args.out_json}")
    elif args.command == "validate":
        validate_cmd(args.json, args.schema)

if __name__ == "__main__":
    main()
